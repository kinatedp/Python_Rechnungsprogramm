import datetime
import locale
import multiprocessing as mp
import os
import platform
import subprocess

import dateparser
import dateutil.parser
from openpyxl import load_workbook


class CRechAuswert_Element:
    rechDate: str
    rechSum: float
    strCustName: str
    strRechNum: str


def checkIfFileIsExcelFile(fPath:str):
    try:
        if os.path.isfile(fPath):
            if fPath.endswith(".xlsx") or fPath.endswith(".xlsm"):
                if not fPath.endswith("_Auswertung.xlsx"):
                    return True

    except Exception as exc:
        print(exc)

    return False


def getRechnungDetailsFromFile(pWorkingDir:str, pFile:str):
    tmpRechAuswerung_Element = CRechAuswert_Element()
    try:
        path2File = os.path.join(pWorkingDir, pFile)

        if checkIfFileIsExcelFile(path2File):
            try:
                currWorkbook = load_workbook(filename=path2File)
                currWorksheet = currWorkbook.active
                maxRow = len(currWorksheet['A'])

                currRow = 12

                while currRow < maxRow:
                    currRow = currRow + 1
                    # wenn zwischensumme vorhanden ist muss der netto wert verwendet werden;
                    # ansonsten ist nettowert = rechnungssume
                    if currWorksheet.cell(row=currRow, column=1).value == "Zwischensumme" or currWorksheet.cell(
                            row=currRow, column=1).value == "Rechnungssumme":
                        tmpRechAuswerung_Element.rechSum = currWorksheet.cell(row=currRow, column=7).value
                        tmpRechAuswerung_Element.strCustName = currWorksheet.cell(row=8, column=2).value
                        tmpRechAuswerung_Element.strRechNum = currWorksheet.cell(row=14, column=6).value

                        try:
                            # noinspection PyUnusedLocal
                            dtRechDatum = None
                            try:
                                # prüfe, ob Wert im Date-Format vorliegt (Datetime gibts nicht)
                                # noinspection PyUnusedLocal
                                tmpRechDate = currWorksheet.cell(row=12, column=6).value.date
                                # falls ja liefert dies ein echtes datetime-obj zurück
                                dtRechDatum = currWorksheet.cell(row=12, column=6).value

                            except Exception as exc:
                                # currentGUI.printExceptionDetails(exc, False)
                                print(exc)
                                # falls nicht haben wir einen deutschen datumsstring
                                strGermanRechDatum = str(currWorksheet.cell(row=12, column=6).value)
                                dtRechDatum = dateparser.parse(strGermanRechDatum, settings={'DATE_ORDER': 'DMY'})

                            # tmpRechAuswerung_Element.rechDate = strRechDatum
                            tmpRechAuswerung_Element.rechDate = dtRechDatum
                        except Exception as exc:
                            print(exc)
                            # currentGUI.printExceptionDetails(exc, True)
                        break

                currWorkbook.close()
                return tmpRechAuswerung_Element

            except Exception as exc:
                print(exc)
                # currentGUI.printExceptionDetails(exc, True)

    except Exception as exc:
        print(exc)
        # currentGUI.printExceptionDetails(exc, True)

    return None


def getRechAuswertElements(pWorkingDir:str):
    lFiles = os.listdir(pWorkingDir)
    lRechAuswertElements = []

    for pFile in lFiles:
        retObj = getRechnungDetailsFromFile(pWorkingDir, pFile)
        lRechAuswertElements.append(retObj)

    return lRechAuswertElements


def PARALLEL_getRechAuswertElements(pWorkingDir:str, currentGUI):
    lFiles = os.listdir(pWorkingDir)
    currentGUI.printInfoToLog("Führe parallele Verarbeitung mit " + str(mp.cpu_count()) + "Kernen aus", True)

    # Step 1: Init multiprocessing.Pool()
    pool = mp.Pool(mp.cpu_count())

    # Step 2: Paralle Ausführung
    lRechAuswertElements = [pool.apply(getRechnungDetailsFromFile, args=(pWorkingDir, pFile)) for pFile in lFiles]

    # Step 3: Don't forget to close
    pool.close()

    return lRechAuswertElements


def writeRechElementsToExcel(lRechAuswertElements, pFileTemplate:str, pFileSave:str, currentGUI):
    try:
        currWorkbook = load_workbook(pFileTemplate)
        currWorksheet = currWorkbook["merged"]

        currRow = 1
        currWorksheet.cell(row=currRow, column=1).value = "Nr."
        currWorksheet.cell(row=currRow, column=2).value = "Kunde"
        currWorksheet.cell(row=currRow, column=3).value = "Rechnungsdatum"
        currWorksheet.cell(row=currRow, column=4).value = "Rechnungssumme"

        currRow = 2

        for RechElement in lRechAuswertElements:
            if RechElement is not None:
                currentGUI.printInfoToLog("Starte mit RechNr. " + str(RechElement.strRechNum), True)
                # erzeugt eine aufsteigende nummer -> ersetzt durch RechNR
                # currWorksheet.cell(row=currRow, column=1).value = str(currRow - 1)
                currWorksheet.cell(row=currRow, column=1).value = str(RechElement.strRechNum)
                currWorksheet.cell(row=currRow, column=2).value = RechElement.strCustName
                currWorksheet.cell(row=currRow, column=4).value = RechElement.rechSum

                tmpDate: datetime.datetime

                if isinstance(RechElement.rechDate, datetime.datetime):
                    tmpDate = RechElement.rechDate

                else:
                    tmpDate = dateutil.parser.parse(RechElement.rechDate)

                locale.setlocale(locale.LC_ALL, "de_DE")
                currWorksheet.cell(row=currRow, column=3).value = tmpDate.strftime("%d.%m.%Y")
                currWorksheet.cell(row=currRow, column=3).number_format = "dd.mm.yyyy"
                currWorksheet.cell(row=currRow, column=5).value = tmpDate.strftime("%B")
                currRow = currRow + 1

        currWorkbook.save(filename=pFileSave)
        currWorkbook.close()

        currentGUI.printInfoToLog("Ende Rechnungsabarbeitung", True)

    except Exception as exc:
        currentGUI.printExceptionDetails(exc, True)


def openExcelWithStandardApp(pWorkingFile:str, currentGUI):
    try:
        if platform.system() == 'Darwin':  # macOS
            subprocess.call(('open', pWorkingFile))
        elif platform.system() == 'Windows':  # Windows
            os.startfile(pWorkingFile)
        else:  # linux variants
            subprocess.call(('xdg-open', pWorkingFile))

    except Exception as exc:
        currentGUI.printExceptionDetails(exc, True)
