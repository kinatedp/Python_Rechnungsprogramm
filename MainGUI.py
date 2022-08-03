import datetime
import locale
import os
import pickle
import platform
import subprocess
from datetime import date
from pathlib import Path
from typing import TextIO

import dateparser
import pandas as pd
from PyQt6 import QtGui, QtWidgets
from PyQt6.QtCore import QLocale
from PyQt6.QtGui import QBrush
from PyQt6.QtWidgets import QMessageBox, QInputDialog, QWidget
from numpy import ndarray
from openpyxl import load_workbook, Workbook, worksheet
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side, BORDER_THICK, BORDER_NONE
from pandas import DataFrame

import RechnungsAuswertung
from HelperClasses import CCustomer, CRechnung, CAbzuege, CArticles
from gui_rechnungsprog import Ui_Dialog

table_boarder_thick_bottom = Border(
    left=Side(border_style=BORDER_NONE, color='00000000'),
    right=Side(border_style=BORDER_NONE, color='00000000'),
    top=Side(border_style=BORDER_NONE, color='00000000'),
    bottom=Side(border_style=BORDER_THICK, color='00000000')
)


# Hinweis: tswPosition kann man über setFlags NICHT auf non-editable setzen
class MainGUI(Ui_Dialog):
    lCustomer: DataFrame = None
    taxRate: float = 0.0
    TEST_ME: bool = False
    strRechNum: str = None
    fPath_SavedExcel: str = None
    currentSelCustomer: CCustomer = None

    def __init__(self, Dialog: QtWidgets.QDialog):
        super().__init__()
        self.setupUi(Dialog)

        self.printInfoToLog("Programm gestartet", True)
        self.loadAndFillConfigPath()
        self.custom_init()


    def loadAndFillConfigPath(self):
        self.printInfoToLog("Lese Config-Pfad aus", True)

        strTmpPath : str = ""

        try:
            strTmpPath = self.loadValueFromFile(self.lFName_PathToConfigDir.text())

        except Exception as exc:
            self.printExceptionDetails(exc, True)

        if strTmpPath == "" or not os.path.isdir(strTmpPath):
            strTmpPath = self.getSelectedDirFromFDialog()

            if strTmpPath == "":
                self.msgBoxWrapper("Kein Ordner ausgewählt.\nBeende Programm")
                exit(-1)

            tmpPath = os.path.normpath(strTmpPath)

            try:
                self.writeValueToFile(tmpPath, self.lFName_PathToConfigDir.text())
            except Exception as exc:
                self.printExceptionDetails(exc, True)
        else:
            tmpPath = os.path.normpath(strTmpPath)

        self.lE_path2ConfigDir.setText(tmpPath)
        self.printInfoToLog("Config-Pfad: " + str(tmpPath), True)


    def fillTableWidgetsBackgrounds(self):
        tmpBrush: QBrush = QBrush(QtGui.QColor(240, 240, 240))
        currRow: int = 0

        while currRow < self.twPositionen.rowCount():
            self.twPositionen.item(currRow, 0).setBackground(tmpBrush)
            self.twPositionen.item(currRow, 1).setBackground(tmpBrush)
            self.twPositionen.item(currRow, 2).setBackground(tmpBrush)
            self.twPositionen.item(currRow, 5).setBackground(tmpBrush)
            currRow += 1

        currRow = 0
        while currRow < self.twAbzuege.rowCount():
            self.twAbzuege.item(currRow, 0).setBackground(tmpBrush)
            self.twAbzuege.item(currRow, 1).setBackground(tmpBrush)
            self.twAbzuege.item(currRow, 2).setBackground(tmpBrush)
            self.twAbzuege.item(currRow, 5).setBackground(tmpBrush)
            currRow += 1

    """
    def fillConfigPathForWinAndMac(self):
        # TODO
        if platform.system() == 'Darwin':  # macOS
            if True:
                tmpPath = self.getSelectedDirFromFDialog()
            else:
                tmpPath = ""

            self.lE_path2ConfigDir.setText(tmpPath)
        elif platform.system() == 'Windows':  # Windows
            self.lE_path2ConfigDir.setText("C:\\Users\\PatrickKinateder\\Documents\\PRIVAT\\RePro")
    """

    def showGUILog(self, bShow: bool):
        if bShow:
            self.pTE_Logger.show()
            self.lLogHeader.show()

        else:
            self.pTE_Logger.hide()
            self.lLogHeader.hide()

    def setLogFlagInGUIFromFile(self):
        self.chkShowLog.setChecked(self.loadShowLogFlagFromFile())

    def custom_init(self):
        # da Sprache auf MAC immer englisch ist, versuch dies auf Deutsch zu forcen
        # in GUI-Datei würde es überschrieben werden
        try:
            self.calendarWidget.setLocale(QLocale(QLocale.Language.German))
        except Exception as exc:
            self.printExceptionDetails(exc, True)

        self.loadPathToDirsToGUI()

        self.setLogFlagInGUIFromFile()
        # passiert bereits durch vorherige methode und dort über change-event
        # self.showGUILog(self.chkShowLog.isChecked())

        self.checkIfNecessaryFilesExist()

        self.tabWidget.setCurrentIndex(0)

        self.currentSelCustomer = CCustomer()

        self.loadAndSetCustomerList()
        self.findAndSetGUIElements()
        self.fill_cbClients()

        self.handleYearChanged(self.checkIfYearChanged())
        self.setAndLoadTaxRate()

        self.resetDialog()

        if self.TEST_ME:
            self.fillFormForTest()

    def setAndLoadTaxRate(self):
        self.taxRate = self.loadTaxRateFromFile()
        self.dsbTaxRate.setValue(self.taxRate)

    def fillFormForTest(self):
        self.tbAuftragsort.setText("Test-Auftragsort")
        self.tbAuftragsnummer.setText("Test-Auftr.Nr")
        self.twPositionen.item(0, 4).setText("11")
        self.twPositionen.item(1, 4).setText("22")
        self.twPositionen.item(2, 4).setText("3,3")
        self.chk_AddTaxes.setChecked(True)

    def checkIfNecessaryFilesExist(self):
        # Template für Rechnungsauswertungen
        if not self.checkIfFileExists(self.lE_path2ConfigDir.text(), self.lFName_Template_RechAuswertung.text(), True):
            self.msgBoxWrapper(
                "Datei " + self.lFName_Template_RechAuswertung.text() + " fehlt.\n\nPfad: " + self.lE_path2ConfigDir.text() + "\n\nBeende Programm.")
            exit(-1)

        # Template für Erstellung der Rechnungen
        if not self.checkIfFileExists(self.lE_path2ConfigDir.text(), self.lFName_Rechnungsmuster.text(), True):
            self.msgBoxWrapper(
                "Datei " + self.lFName_Rechnungsmuster.text() + " fehlt.\n\nPfad: " + self.lE_path2ConfigDir.text() + "\n\nBeende Programm.")
            exit(-1)

        # Kunden-Datei fehlt
        if not self.checkIfFileExists(self.lE_path2ConfigDir.text(), self.lFName_Customer.text(), True):
            self.msgBoxWrapper(
                "Datei " + self.lFName_Customer.text() + " fehlt.\n\nPfad: " + self.lE_path2ConfigDir.text() + "\n\nBeende Programm.")
            exit(-1)

        # Datei für Positionen fehlt => Ohne kann Programm NICHT fortgesetzt werden
        if not self.checkIfFileExists(self.lE_path2ConfigDir.text(), self.lFName_Positionen.text(), True):
            self.msgBoxWrapper(
                "Datei " + self.lFName_Positionen.text() + " fehlt.\n\nPfad: " + self.lE_path2ConfigDir.text() + "\n\nBeende Programm.")
            exit(-1)

        # Datei für Abzüge fehlt -> ohne KANN Programm fortgesetzt werden
        if not self.checkIfFileExists(self.lE_path2ConfigDir.text(), self.lFName_Abzuege.text(), True):
            retOpt = self.msgBoxYesNoWrapper(
                "Die Datei für die Abzüge ist nicht vorhanden. \n\nPfad: " + self.lE_path2ConfigDir.text() + "\n\nProgramm beenden?",
                "Abzüge nicht vorhanden")
            if retOpt == QtWidgets.QMessageBox.StandardButton.Yes:
                exit(-1)

        # Letztes Rechnungsjahr
        if not self.checkIfFileExists(self.lE_path2ConfigDir.text(), self.lFName_LastRechYear.text(), True):
            inpRechJahr: int
            okSelected: bool

            try:
                iYearTemplate: int = datetime.date.today().year
            except:
                iYearTemplate: int = 2020

            inpRechJahr, okSelected = QInputDialog().getInt(QWidget(), "Aktuelles Rechnungsjahr noch nicht vorhanden",
                                                            "Bitte aktuelles Jahr angeben (z.B. 2022): ", iYearTemplate,
                                                            2000, 2100)
            if okSelected:
                self.writeValueToFile(str(inpRechJahr),
                                      os.path.join(self.lE_path2ConfigDir.text(), self.lFName_LastRechYear.text()))

    def checkIfFileExists(self, pDirectory: str, pFileName: str, bPrintMsg: bool):
        pathToCheck: str = os.path.join(os.path.normpath(pDirectory), pFileName)

        if os.path.exists(pathToCheck):
            self.printInfoToLog("Datei " + pFileName + " vorhanden", bPrintMsg)
            return True

        self.printInfoToLog("Datei " + pFileName + " NICHT vorhanden", bPrintMsg)
        return False

    def fill_cbClients(self):
        if self.cbClients.count() > 0:
            self.cbClients.clear()

        # nicht mit addItems damit es alphabetisch sortiert ist
        if self.lCustomer is not None:
            df_array: ndarray = self.lCustomer.values
            for row in range(self.lCustomer.shape[0]):
                self.cbClients.addItem(df_array[row, 0])

        self.cbClients.setCurrentText("Femo GmbH")

    def loadLatestRechNumber(self):
        fPath: str = os.path.join(os.path.normpath(self.lE_path2ConfigDir.text()), self.lFName_RechNr.text())
        try:
            fObj: TextIO = open(fPath, "r")
            self.strRechNum = fObj.read()
            fObj.close()
            self.printInfoToLog("Rechnungsnr. " + self.strRechNum + " ausgelesen", True)

        except Exception as exc:
            self.printExceptionDetails(exc, True)
            self.strRechNum = "1"
            self.printInfoToLog("Rechnungsnr konnte NICHT ausgelesen werden. Wird auf 1 gesetzt", True)
            self.resetRechNumToOneInFile()

        self.setRechNumInGUI()

    def loadLatestYearFromFile(self):
        fPath: str = os.path.join(os.path.normpath(self.lE_path2ConfigDir.text()), self.lFName_LastRechYear.text())
        try:
            fObj: TextIO = open(fPath, "r")
            strLatestRechYear: str = fObj.read()
            fObj.close()
            self.printInfoToLog("Letztes RechJahr: " + strLatestRechYear, True)

        except Exception as exc:
            self.printExceptionDetails(exc, True)
            strLatestRechYear = str(datetime.date.today().year)

        return strLatestRechYear

    # TODO: Type
    def loadShowLogFlagFromFile(self):
        fPath: str = os.path.join(os.path.normpath(self.lE_path2ConfigDir.text()), self.lFName_ShowLog.text())
        try:
            fObj: TextIO = open(fPath, "r")
            strShowLogFileFlag = fObj.read()
            fObj.close()

        except Exception as exc:
            self.printExceptionDetails(exc, True)
            strShowLogFileFlag = "Yes"
            # Muss nicht mehr erfolgen => wird durch Change event geschrieben
            # self.writeValueToFile(strShowLogFileFlag, fPath)

        return strShowLogFileFlag == "True"

    def loadTaxRateFromFile(self):
        dTaxRate = 0
        fPath: str = os.path.join(os.path.normpath(self.lE_path2ConfigDir.text()), self.lFName_UstSatz.text())
        strRetValFromFile: str = self.loadValueFromFile(fPath)

        if strRetValFromFile == "":
            inpTaxRate: float
            okSelected: bool
            # file not found or no value specified -> ask user
            inpTaxRate, okSelected = QInputDialog().getDouble(QWidget(), "Ust.-Satz noch nicht vorhanden",
                                                              "Ust.-Satz (z.B: 19 für 19%) angeben:", 19, 0, 100)
            if okSelected:
                self.writeTaxRateToFile(inpTaxRate)
                strRetValFromFile = str(inpTaxRate)

        try:
            dTaxRate = float(strRetValFromFile)
            self.printInfoToLog("Steuersatz ausgelesen: " + str(dTaxRate) + "%", True)

        except Exception as exc:
            self.printExceptionDetails(exc, True)

        return dTaxRate

    def loadValueFromFile(self, path: str):
        strRet: str = ""
        try:
            fObj: TextIO = open(path, "r")
            strRet = fObj.read()
            fObj.close()

        except Exception as exc:
            self.printExceptionDetails(exc, True)

        return strRet

    def loadPathToDirsToGUI(self):
        # Entfernt; Verwende nur noch Hardcoded Pfad aus GUI sonst Probleme beim Start
        # fPath2ConfigDir :str = os.path.join(os.path.normpath(self.lE_path2ConfigDir.text()), self.lFName_PathToConfigDir.text())

        fPath2OutputDir: str = os.path.join(os.path.normpath(self.lE_path2ConfigDir.text()),
                                            self.lFName_PathToRechOutputDir.text())

        try:
            self.lE_path2Output_rechnungen.setText(self.loadValueFromFile(fPath2OutputDir))
            self.enableBtnErstelleRechnung()

            if self.lE_path2Output_rechnungen.text() != "":
                self.printInfoToLog("RechOutput Ordner ausgelesen: " + self.lE_path2Output_rechnungen.text(), True)

        except Exception as exc:
            self.printExceptionDetails(exc, True)

        # Entfernt; Verwende nur noch Hardcoded Pfad aus GUI sonst Probleme beim Start
        # self.lE_path2ConfigDir.setText(self.loadValueFromFile(fPath2ConfigDir))
        self.lE_pathRechnungen_auswert.setText(self.lE_path2Output_rechnungen.text())
        self.on_pathRechnungen_auswert_changed()

    def setRechNumInGUI(self):
        rechYear: str = str((dateparser.parse(self.tbRechDatum.text(), settings={'DATE_ORDER': 'DMY'})).year) + "-"

        if len(self.strRechNum) == 3:
            self.tbRechNummer.setText(rechYear + self.strRechNum)

        elif len(self.strRechNum) == 2:
            self.tbRechNummer.setText(rechYear + "0" + self.strRechNum)

        else:
            self.tbRechNummer.setText(rechYear + "00" + self.strRechNum)

        try:
            self.sbRechNumToSet.setValue(int(self.strRechNum))

        except Exception as exc:
            self.printExceptionDetails(exc, True)
            self.sbRechNumToSet.setValue(1)

    def cbCustomer_selectionChanged(self):
        if self.cbClients.count() > 0:
            # set customer street + city
            df_array: ndarray = self.lCustomer.values
            for row in range(self.lCustomer.shape[0]):
                if df_array[row, 0] == self.cbClients.currentText():
                    self.currentSelCustomer.cust_name = self.cbClients.currentText()
                    self.currentSelCustomer.cust_street = df_array[row, 1]
                    self.currentSelCustomer.cust_city = df_array[row, 2]

                    self.lCustomer_street.setText(self.currentSelCustomer.cust_street)
                    self.lCustomer_city.setText(self.currentSelCustomer.cust_city)
                    break
            self.printInfoToLog("Kunde geändert zu: " + self.currentSelCustomer.cust_name, True)
            self.loaddataWithCSV()

    def loadAndSetCustomerList(self):
        pathToFile: str = os.path.join(os.path.normpath(self.lE_path2ConfigDir.text()), self.lFName_Customer.text())
        try:
            self.lCustomer = pd.read_csv(pathToFile, sep=";")
            self.printInfoToLog("Anzahl Kunden aus Datei: " + str(len(self.lCustomer)), True)

        except Exception as exc:
            self.printExceptionDetails(exc, True)
            self.printInfoToLog((self.lFName_Customer.text() + " not found"), True)

    def on_selected_date_changed(self):
        self.tbAuftragsdatum.setDate(self.calendarWidget.selectedDate())

    def findAndSetGUIElements(self):
        wPosNr: int = 128
        wPosBez: int = 360
        wPosEinheit: int = 65
        wPosPreis: int = 68
        wPosMenge: int = 95
        wPosGesamt: int = 100

        self.cbClients.currentIndexChanged.connect(self.cbCustomer_selectionChanged)
        self.tbRechNummer.textChanged.connect(self.tbRechNum_text_changed)

        self.btnSave.clicked.connect(self.on_click_erstelleRechnung)

        self.calendarWidget.selectionChanged.connect(self.on_selected_date_changed)

        self.twPositionen.setColumnWidth(0, wPosNr)
        self.twPositionen.setColumnWidth(1, wPosBez)
        self.twPositionen.setColumnWidth(2, wPosEinheit)
        self.twPositionen.setColumnWidth(3, wPosPreis)
        self.twPositionen.setColumnWidth(4, wPosMenge)
        self.twPositionen.setColumnWidth(5, wPosGesamt)

        self.twAbzuege.setColumnWidth(0, wPosNr)
        self.twAbzuege.setColumnWidth(1, wPosBez)
        self.twAbzuege.setColumnWidth(2, wPosEinheit)
        self.twAbzuege.setColumnWidth(3, wPosPreis)
        self.twAbzuege.setColumnWidth(4, wPosMenge)
        self.twAbzuege.setColumnWidth(5, wPosGesamt)

        self.btnSelDirConfig.clicked.connect(self.on_click_SelDirConfig)
        self.btnSelDirOutputRech.clicked.connect(self.on_click_SelDirOutputRech)
        self.btnSelDirRechAuswert.clicked.connect(self.on_click_SelDirRechAuswert)
        self.btn_ErstelleAuswertung.clicked.connect(self.on_click_ErstelleAuswertung)

        self.twPositionen.cellChanged.connect(self.on_cellChanged_event_pos)
        self.twAbzuege.cellChanged.connect(self.on_cellChanged_event_abz)

        self.tbAuftragsnummer.textChanged.connect(self.on_necessaryTextboxes_changed)
        self.tbAuftragsort.textChanged.connect(self.on_necessaryTextboxes_changed)
        self.btnResetRechNumToOne.clicked.connect(self.on_click_resetRechNum)
        self.btnChangeRechNum.clicked.connect(self.on_click_changeRechNumFile)
        self.tbAuftragsdatum.dateChanged.connect(self.on_tbAuftragsdatum_dateChanged)
        self.lE_pathRechnungen_auswert.textChanged.connect(self.on_pathRechnungen_auswert_changed)

        self.btnOpenAbzWithApp.clicked.connect(self.on_click_OpenAbzWithApp)
        self.btnOpenPosWithApp.clicked.connect(self.on_click_OpenPosWithApp)

        self.btnResetGUI.clicked.connect(self.resetDialog)
        self.chkShowLog.stateChanged.connect(self.on_ShowLog_changed)
        self.pTE_Logger.textChanged.connect(self.on_loggerText_changed)

        self.dsbTaxRate.textChanged.connect(self.on_taxRateText_changed)

    def on_taxRateText_changed(self):
        self.taxRate = 0.0

        try:
            self.taxRate = self.dsbTaxRate.value()

        except Exception as exc:
            self.printExceptionDetails(exc, True)

        self.chk_AddTaxes.setTitle("Erstelle mit " + str(self.taxRate) + "% Ust.")
        self.writeTaxRateToFile(self.taxRate)
        self.calcCurrentRechSum()

    def on_loggerText_changed(self):
        self.pTE_Logger.horizontalScrollBar().setValue(0)

    def on_ShowLog_changed(self):
        self.showGUILog(self.chkShowLog.isChecked())
        self.printInfoToLog("Anzeige des Logs geändert zu: " + str(self.chkShowLog.isChecked()), True)
        self.writeShowLogFlagToFile()

    def writeShowLogFlagToFile(self):
        fPath: str = os.path.join(os.path.normpath(self.lE_path2ConfigDir.text()), self.lFName_ShowLog.text())
        self.writeValueToFile(str(self.chkShowLog.isChecked()), fPath)

    def on_click_OpenAbzWithApp(self):
        fPath: str = os.path.join(os.path.normpath(self.lE_path2ConfigDir.text()), self.lFName_Abzuege.text())
        self.openFileWithStdApp(fPath)

    def on_click_OpenPosWithApp(self):
        fPath: str = os.path.join(os.path.normpath(self.lE_path2ConfigDir.text()), self.lFName_Positionen.text())
        self.openFileWithStdApp(fPath)

    def createIfNotExisting(self, fPath: str):
        retOpt: int = self.msgBoxYesNoWrapper(
            "Datei ist nicht vorhanden: \n" + fPath + "\n\n Wollen Sie die Datei anlegen?",
            "Fehler beim Öffnen")

        if retOpt == QtWidgets.QMessageBox.StandardButton.Yes:
            try:
                myFile: Path = Path(fPath)
                myFile.touch(exist_ok=True)
            except Exception as exc:
                self.printExceptionDetails(exc, True)

    def openFileWithStdApp(self, fPath: str):
        self.printInfoToLog("Versuche " + str(fPath) + " mit Standardanwendung zu Öffnen...", True)
        try:
            if platform.system() == 'Darwin':  # macOS
                subprocess.call(('open', fPath))
            elif platform.system() == 'Windows':  # Windows
                os.startfile(fPath)
            else:  # linux variants
                subprocess.call(('xdg-open', fPath))

        except Exception as exc:
            self.printExceptionDetails(exc, True)
            self.createIfNotExisting(fPath)

    # TODO: Type
    def on_click_ErstelleAuswertung(self):
        if self.lE_pathRechnungen_auswert.text() != "":
            try:
                pFileTemplateAuswertungen: str = os.path.join(os.path.normpath(self.lE_path2ConfigDir.text()),
                                                              self.lFName_Template_RechAuswertung.text())
                workingDir = os.path.normpath(self.lE_pathRechnungen_auswert.text())
                strCurrDate = date.today().strftime("%Y-%m-%d")
                pFileSave_fName = strCurrDate + "_Auswertung.xlsx"
                pFileSave: str = os.path.join(workingDir, pFileSave_fName)

                # wird nicht verwendet -> höhere Laufzeit als normaler Modus
                # lRechAuswertElements = RechnungsAuswertung.PARALLEL_getRechAuswertElements(workingDir, self)
                lRechAuswertElements = RechnungsAuswertung.getRechAuswertElements(workingDir)

                if len(lRechAuswertElements) == 0:
                    self.printInfoToLog("Keine Rechnungen gefunden - Auswertung nicht möglich", True)

                else:
                    self.printInfoToLog("Erstelle Auswertung mit " + str(len(lRechAuswertElements)) + " Rechnungen",
                                        True)

                    RechnungsAuswertung.writeRechElementsToExcel(lRechAuswertElements, pFileTemplateAuswertungen,
                                                                 pFileSave, self)

                    if self.checkIfFileExists(workingDir, pFileSave_fName, False):
                        self.printInfoToLog("Datei erstellt: " + pFileSave_fName + "\tin Ordner: " + workingDir, True)
                        self.printInfoToLog("Öffne Auswertung mit Excel", True)
                        RechnungsAuswertung.openExcelWithStandardApp(pFileSave, self)

                    else:
                        self.printInfoToLog("Fehler beim Erstellen der Auswertung", True)

            except Exception as exc:
                self.printExceptionDetails(exc, True)

    def on_pathRechnungen_auswert_changed(self):
        tmpPath: str = self.lE_pathRechnungen_auswert.text()
        fCnt: int = 0

        if tmpPath is not None and tmpPath != "":
            if os.path.isdir(tmpPath):
                lElementsInDir: list = os.listdir(tmpPath)
                for elem in lElementsInDir:
                    if RechnungsAuswertung.checkIfFileIsExcelFile(os.path.join(tmpPath, elem)):
                        fCnt += 1
                self.lCnt_FilesForAuswert.setText(str(fCnt))

        self.btn_ErstelleAuswertung.setEnabled(fCnt != 0)

        if fCnt == 0:
            self.printInfoToLog("Pfad für Rechnungsauswertung geändert - 0 Rechnungen gefunden", True)
        else:
            self.printInfoToLog("Pfad für Rechnungsauswertung geändert - " + str(fCnt) + " Rechnungen gefunden", True)

    def on_tbAuftragsdatum_dateChanged(self):
        self.calendarWidget.setSelectedDate(self.tbAuftragsdatum.date())

    def on_necessaryTextboxes_changed(self):
        self.enableBtnErstelleRechnung()

    def on_cellChanged_event_pos(self, row: int, column: int):
        if column == 4 or column == 3:
            tmpMenge: float = 0.0
            tmpPrice: float = 0.0

            try:
                if self.twPositionen.item(row, 4) is not None:
                    tmpMenge = float(self.twPositionen.item(row, 4).text().replace(",", "."))

                if self.twPositionen.item(row, 3) is not None:
                    tmpPrice = float(self.twPositionen.item(row, 3).text().replace(",", "."))

            except Exception as exc:
                self.printExceptionDetails(exc, False)

            tmpGesamtpreis = str(round(tmpPrice * tmpMenge, 2))
            try:
                if self.twPositionen.item(row, 5) is not None:
                    self.twPositionen.item(row, 5).setText(tmpGesamtpreis)
            except Exception as exc:
                self.printExceptionDetails(exc, False)
            self.calcCurrentRechSum()

    def on_cellChanged_event_abz(self, row: int, column: int):
        if column == 4 or column == 3:
            tmpMenge: float = 0.0
            tmpPrice: float = 0.0

            try:
                if self.twAbzuege.item(row, 4) is not None:
                    tmpMenge = float(self.twAbzuege.item(row, 4).text().replace(",", "."))

                if self.twAbzuege.item(row, 3) is not None:
                    tmpPrice = float(self.twAbzuege.item(row, 3).text().replace(",", "."))

            except Exception as exc:
                self.printExceptionDetails(exc, False)

            tmpGesamtpreis = str(round(tmpPrice * tmpMenge, 2))
            try:
                if self.twAbzuege.item(row, 5) is not None:
                    self.twAbzuege.item(row, 5).setText(tmpGesamtpreis)
            except Exception as exc:
                self.printExceptionDetails(exc, False)
            self.calcCurrentRechSum()

    def calcCurrentRechSum(self):
        currRow: int = 0
        currRechSum: float = 0.0

        while currRow < self.twPositionen.rowCount():
            try:
                if self.twPositionen.item(currRow, 5) is not None:
                    tmpLineSum: float = float(self.twPositionen.item(currRow, 5).text())
                    currRechSum += tmpLineSum

            except Exception as exc:
                self.printExceptionDetails(exc, False)
            currRow += 1

        currRow = 0
        while currRow < self.twAbzuege.rowCount():
            try:
                if self.twAbzuege.item(currRow, 5) is not None:
                    tmpLineSum: float = float(self.twAbzuege.item(currRow, 5).text())
                    currRechSum -= tmpLineSum
            except Exception as exc:
                self.printExceptionDetails(exc, False)
            currRow += 1

        self.teCurrRechSum.setText(str(round(currRechSum, 2)))

        ustBetrag: float = round(currRechSum * self.taxRate / 100, 2)
        self.teCurrRechTax.setText(str(ustBetrag))
        self.teCurrRechSumWithTax.setText(str(ustBetrag + currRechSum))

        self.enableBtnErstelleRechnung()

    def checkRechOutputDirIsGiven(self):
        if self.lE_path2Output_rechnungen.text() != "":
            return os.path.isdir(self.lE_path2Output_rechnungen.text())
        return False

    # TODO
    def enableBtnErstelleRechnung(self):
        currRechnung: CRechnung = self.getCurrentRechnungDetails()
        self.lMsgNoRechOutputDir.hide()

        if self.checkNecessaryDetailsRechnung(
                currRechnung) and self.checkCurrentRechSumNotNullOrEmpty() and self.checkRechOutputDirIsGiven():
            self.btnSave.setEnabled(True)
        else:
            if not self.checkRechOutputDirIsGiven():
                self.lMsgNoRechOutputDir.show()
            self.btnSave.setEnabled(False)

    def checkCurrentRechSumNotNullOrEmpty(self):
        try:
            tmpFloatSum: float = float(self.teCurrRechSum.text())
            if tmpFloatSum != 0:
                return True

        except Exception as exc:
            self.printExceptionDetails(exc, True)

        return False

    @staticmethod
    def getSelectedDirFromFDialog():
        fDiag: QtWidgets.QFileDialog = QtWidgets.QFileDialog()
        fDiag.setFileMode(QtWidgets.QFileDialog.FileMode.Directory)
        dirPath: str = fDiag.getExistingDirectory()

        return dirPath

    def on_click_SelDirConfig(self):
        tmpSelDir: str = self.getSelectedDirFromFDialog()
        if tmpSelDir != "":
            if os.path.isdir(tmpSelDir):
                self.lE_path2ConfigDir.setText(tmpSelDir)
                self.writePathToConfigDirToFile()

    def on_click_SelDirOutputRech(self):
        tmpSelDir: str = self.getSelectedDirFromFDialog()
        if tmpSelDir != "":
            if os.path.isdir(tmpSelDir):
                self.lE_path2Output_rechnungen.setText(tmpSelDir)
                self.printInfoToLog("Pfad zum Speichern von Rechnungen geändert zu: " + str(tmpSelDir), True)
                self.writePathToRechOutputDirToFile()
                self.lE_pathRechnungen_auswert.setText(tmpSelDir)
                self.enableBtnErstelleRechnung()

    def on_click_SelDirRechAuswert(self):
        tmpSelDir: str = self.getSelectedDirFromFDialog()
        if tmpSelDir != "":
            if os.path.isdir(tmpSelDir):
                self.lE_pathRechnungen_auswert.setText(tmpSelDir)

    def tbRechNum_text_changed(self):
        tmpRechNum: str = self.tbRechNummer.text()[self.tbRechNummer.text().find("-") + 1:len(self.tbRechNummer.text())]
        self.strRechNum = tmpRechNum

    def resetDialog(self):
        self.tbRechDatum.setDate(date.today())
        self.tbAuftragsdatum.setDate(date.today())
        self.tbAuftragsort.clear()
        self.tbAuftragsnummer.clear()
        self.tbRechNummer.clear()
        self.calendarWidget.setSelectedDate(date.today())

        self.calendarWidget.selectedDate()
        self.loadLatestRechNumber()

        self.loaddataWithCSV()
        self.teCurrRechSum.setText("0.0")

        self.teCurrRechTax.setText("0.0")
        self.teCurrRechSumWithTax.setText("0.0")

        self.btnSave.setEnabled(False)

    def on_click_erstelleRechnung(self):
        self.printInfoToLog("Erstelle Rechnung", True)
        self.fPath_SavedExcel = ""  # wird beim speichern gesetzt
        was_success: bool = self.writeRechnungToStorage()

        if was_success:
            self.printInfoToLog("Rechnung erstellt: " + self.fPath_SavedExcel, True)
            self.actionsAfterRechnungSave()
            self.printInfoToLog("Öffne Rechnung mit Excel", True)
            self.openExcelFileAfterwards()

    def openExcelFileAfterwards(self):
        if self.chk_OpenFileAfterCreation.isChecked():
            if os.path.isfile(self.fPath_SavedExcel):
                try:
                    if platform.system() == 'Darwin':  # macOS
                        subprocess.call(('open', self.fPath_SavedExcel))
                    elif platform.system() == 'Windows':  # Windows
                        os.startfile(self.fPath_SavedExcel)
                    else:  # linux variants
                        subprocess.call(('xdg-open', self.fPath_SavedExcel))
                except Exception as exc:
                    self.printExceptionDetails(exc, True)

    @staticmethod
    def checkNecessaryDetailsRechnung(currentRechnung: CRechnung):
        if currentRechnung.strRechnungsNum == "" or currentRechnung.strRechnungsdatum == "" or currentRechnung.strAuftragsort == "" or currentRechnung.strAuftragsdatum == "" or currentRechnung.strAuftragsnummer == "":
            return False

        else:
            return True

    def loaddataWithCSV(self):
        self.loadArticles()
        self.loadAbzuege()

        self.fillTableWidgetsBackgrounds()

    def loadArticles(self):
        self.twPositionen.clear()
        self.twPositionen.clearContents()

        lTableLabels = ["Pos.Nr.", "Position", "Einheit", "Preis", "Menge", "Gesamt"]
        self.twPositionen.setHorizontalHeaderLabels(lTableLabels)

        try:
            fName: str = self.lFName_Positionen.text()
            fPath: str = os.path.join(os.path.normpath(self.lE_path2ConfigDir.text()), fName)

            if os.path.isfile(fPath):
                tmpUseCols = ["Positionsnummer", "Position", "Einheit", self.cbClients.currentText(), "Menge",
                              "Gesamtpreis"]
                dfData: DataFrame = pd.read_csv(fPath, sep=";", usecols=tmpUseCols)
                dfData = dfData[dfData[
                    self.cbClients.currentText()].notna()]  # entferne alle Zeilen bei denen die Preis-Spalte nicht befüllt ist

                self.write_df_to_qtable(dfData, self.twPositionen)
                self.printInfoToLog("Positionen ausgelesen: " + str(dfData.size / 6).replace(".0", ""), True)

        except Exception as exc:
            self.printExceptionDetails(exc, True)

        self.twPositionen.setHorizontalHeaderLabels(lTableLabels)

    def loadArticles_OLD(self):
        self.twPositionen.clear()
        self.twPositionen.clearContents()

        lTableLabels = ["Positionsnummer", "Position", "Einheit", "Preis", "Menge", "Gesamtpreis"]
        self.twPositionen.setHorizontalHeaderLabels(lTableLabels)

        try:
            fName = "Positionen_" + self.cbClients.currentText() + ".csv"
            fPath: str = os.path.join(os.path.normpath(self.lE_path2ConfigDir.text()), fName)

            if os.path.isfile(fPath):
                dfData: DataFrame = pd.read_csv(fPath, sep=";")

                self.write_df_to_qtable(dfData, self.twPositionen)

        except Exception as exc:
            self.printExceptionDetails(exc, True)

    def loadAbzuege(self):
        self.twAbzuege.clear()
        self.twAbzuege.clearContents()
        lTableLabels = ["AbzugNr.", "Abzug", "Einheit", "Preis", "Menge", "Gesamt"]
        self.twAbzuege.setHorizontalHeaderLabels(lTableLabels)

        try:
            fName = self.lFName_Abzuege.text()
            fPath: str = os.path.join(os.path.normpath(self.lE_path2ConfigDir.text()), fName)

            tmpUseCols = ["Abzugsnummer", "Abzüge", "Einheit", self.cbClients.currentText(), "Menge", "Gesamtpreis"]
            dfData: DataFrame = pd.read_csv(fPath, sep=";", usecols=tmpUseCols)
            dfData = dfData[dfData[
                self.cbClients.currentText()].notna()]  # entferne alle Zeilen bei denen die Preis-Spalte nicht befüllt ist

            self.write_df_to_qtable(dfData, self.twAbzuege)
            self.printInfoToLog("Abzüge ausgelesen: " + str(dfData.size / 6).replace(".0", ""), True)

        except Exception as exc:
            self.printExceptionDetails(exc, True)

        self.twAbzuege.setHorizontalHeaderLabels(lTableLabels)

    def loadAbzuege_OLD(self):
        self.twAbzuege.clear()
        self.twAbzuege.clearContents()
        lTableLabels = ["Abzugsnummer", "Abzug", "Einheit", "Preis", "Menge", "Gesamtabzug"]
        self.twAbzuege.setHorizontalHeaderLabels(lTableLabels)

        try:
            fName: str = "Abzuege_" + self.cbClients.currentText() + ".csv"
            fPath: str = os.path.join(os.path.normpath(self.lE_path2ConfigDir.text()), fName)
            dfData: DataFrame = pd.read_csv(fPath, sep=";")

            self.write_df_to_qtable(dfData, self.twAbzuege)
        except Exception as exc:
            self.printExceptionDetails(exc, True)

    @staticmethod
    # TODO typesf
    def write_df_to_qtable(df, table):
        headers = list(df)
        table.setRowCount(df.shape[0])
        table.setColumnCount(df.shape[1])
        table.setHorizontalHeaderLabels(headers)

        # getting data from df is computationally costly so convert it to array first
        df_array = df.values
        for row in range(df.shape[0]):
            for col in range(df.shape[1]):
                if str(df_array[row, col]) == "nan":
                    table.setItem(row, col, QtWidgets.QTableWidgetItem(str(0)))
                else:
                    table.setItem(row, col, QtWidgets.QTableWidgetItem(str(df_array[row, col])))

    def actionsOnClose(self):
        self.printInfoToLog("Actions on close", True)
        self.writeCurrentRechNumToFile()

    def writeValueToFile(self, value: str, path: str):
        try:
            fObj: TextIO = open(path, "w")
            fObj.write(value)
            fObj.close()

            self.printInfoToLog("Wert: " + str(value) + " in Datei " + str(path) + " geschrieben.", True)

        except Exception as exc:
            self.printExceptionDetails(exc, True)

    def on_click_resetRechNum(self):
        self.resetRechNumToOneInFile()
        # self.resetDialog()
        self.msgBoxWrapper("Rechnungsnummer auf 1 gesetzt")
        self.loadLatestRechNumber()

    def on_click_changeRechNumFile(self):
        tmpRechNumToSet: str = str(self.sbRechNumToSet.value())
        self.writeRechNumToFile(tmpRechNumToSet)
        # self.resetDialog()
        self.msgBoxWrapper("Rechnungsnummer auf " + tmpRechNumToSet + " gesetzt")
        self.loadLatestRechNumber()

    def resetRechNumToOneInFile(self):
        self.writeRechNumToFile("1")

    def writeCurrentRechNumToFile(self):
        self.writeRechNumToFile(self.strRechNum)

    def writeRechNumToFile(self, rechNum):
        try:
            fPath: str = os.path.join(os.path.normpath(self.lE_path2ConfigDir.text()), self.lFName_RechNr.text())
            self.writeValueToFile(rechNum, fPath)
            self.printInfoToLog("Rechnungsnr. " + rechNum + " in Datei geschrieben.", True)

        except Exception as exc:
            self.printExceptionDetails(exc, True)

    def writeTaxRateToFile(self, dTaxRate: float):
        fPath: str = os.path.join(os.path.normpath(self.lE_path2ConfigDir.text()), self.lFName_UstSatz.text())
        self.writeValueToFile(str(dTaxRate), fPath)

    def writePathToConfigDirToFile(self):
        fPath: str = os.path.join(os.path.normpath(self.lE_path2ConfigDir.text()), self.lFName_PathToConfigDir.text())
        self.writeValueToFile(os.path.normpath(self.lE_path2ConfigDir.text()), fPath)

    def writePathToRechOutputDirToFile(self):
        fPath: str = os.path.join(os.path.normpath(self.lE_path2ConfigDir.text()),
                                  self.lFName_PathToRechOutputDir.text())
        self.writeValueToFile(self.lE_path2Output_rechnungen.text(), fPath)

    def writeRechnungToStorage(self):
        currRechnung: CRechnung = self.getCurrentRechnungDetails()

        all_neccessary_details_filled: bool = self.checkNecessaryDetailsRechnung(currRechnung)

        if all_neccessary_details_filled:
            fNameRech = self.setupRechFileName(currRechnung)
            fPath: str = os.path.join(os.path.normpath(self.lE_path2Output_rechnungen.text()), fNameRech)
            return self.writeDataToExcel(currRechnung, fPath)

        else:
            self.msgBoxWrapper("Zur Erstellung der Rechnung fehlen noch Informationen")

        return False

    @staticmethod
    def msgBoxWrapper(strMsg: str):
        msgBox: QMessageBox = QMessageBox()
        msgBox.setText(strMsg)
        msgBox.setWindowTitle("Hinweis")
        # msgBox.setWindowIcon(QtGui.QIcon('icon.png'))
        msgBox.setWindowIcon(QtGui.QIcon('icon.ico'))
        msgBox.exec()

    @staticmethod
    def msgBoxYesNoWrapper(strMsg: str, strTitel: str):
        msgBox: QMessageBox = QMessageBox()
        msgBox.setText(strMsg)
        msgBox.setWindowTitle(strTitel)
        # msgBox.setWindowIcon(QtGui.QIcon('icon.png'))
        msgBox.setWindowIcon(QtGui.QIcon('icon.ico'))
        msgBox.setStandardButtons(QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No)
        msgBox.setDefaultButton(QtWidgets.QMessageBox.StandardButton.Yes)
        retVal = msgBox.exec()

        return retVal

    @staticmethod
    def setupRechFileName(currRechnung: CRechnung):
        # Obsolete
        # dRechDatum = dateutil.parser.parse(currRechnung.strRechnungsdatum)
        # strMonth = str(dRechDatum.month)
        # strDay = str(dRechDatum.day)

        # if dRechDatum.month < 10:
        #    strMonth = "0" + str(dRechDatum.month)

        # if dRechDatum.day < 10:
        #    strDay = "0" + str(dRechDatum.day)

        # F-Name soll Jahr,  Rechnungsnummer und Kunde beinhalten und nicht das RechDatum
        # fName_obsolet = str(dRechDatum.year) + "-" + strMonth + "-" + strDay + "_" + currRechnung.strRechnungsNum + "_" + currRechnung.customer.cust_name + ".xlsx"

        fName: str = currRechnung.strRechnungsNum.replace("-", "") + currRechnung.customer.cust_name + ".xlsx"
        return fName

    def handleYearChanged(self, bYearChanged: bool):
        if bYearChanged:
            retOption = self.msgBoxYesNoWrapper(
                "Das Jahr hat sich seit der letzten Rechnung geändert. Wollen Sie die Rechnungsnummer zurücksetzen?",
                "Rechnungsjahr hat sich geändert")

            if retOption == QtWidgets.QMessageBox.StandardButton.Yes:
                self.resetRechNumToOneInFile()

    def saveCurrRechObjToFile(self, currRechnung: CRechnung):
        fPath: str = os.path.join(os.path.normpath(self.lE_path2ConfigDir.text()), "latestCurrRechObj.obj")
        fileToWrite = None

        try:
            fileToWrite = open(fPath, "wb")
            pickle.dump(currRechnung, fileToWrite)
            fileToWrite.close()
            return True

        except Exception as exc:
            if fileToWrite is not None:
                if not fileToWrite.closed:
                    fileToWrite.close()
            self.printExceptionDetails(exc, True)
            self.printInfoToLog("Fehler beim Schreiben des Rechnungsobjekt-Files", True)

        return False

    def writeDataToExcel(self, currentRech: CRechnung, fPathSaveLocation: str):
        try:
            fPathMusterFile: str = os.path.join(os.path.normpath(self.lE_path2ConfigDir.text()),
                                                self.lFName_Rechnungsmuster.text())

            workbook: Workbook = load_workbook(filename=fPathMusterFile)
            actSheet: worksheet = workbook["Einstieg"]

            actSheet["B8"].value = currentRech.customer.cust_name
            actSheet["B9"].value = currentRech.customer.cust_street
            actSheet["B10"].value = currentRech.customer.cust_city

            actSheet["F12"].value = currentRech.strRechnungsdatum
            actSheet["F14"].value = currentRech.strRechnungsNum
            actSheet["A16"].value = currentRech.strAuftragsort
            actSheet["F16"].value = currentRech.strAuftragsnummer
            actSheet["F17"].value = currentRech.strAuftragsdatum

            startRowForPositions: int = 20
            zwischensumme: float = 0.0

            # Positionen ausgeben
            for pos in currentRech.lArticles:
                actSheet.cell(row=startRowForPositions, column=2).value = pos.art_id
                actSheet.cell(row=startRowForPositions, column=3).value = pos.art_name
                actSheet.cell(row=startRowForPositions, column=4).value = pos.art_unit
                actSheet.cell(row=startRowForPositions, column=5).value = pos.art_price
                actSheet.cell(row=startRowForPositions, column=6).value = pos.art_amount
                actSheet.cell(row=startRowForPositions, column=7).value = pos.art_total_price

                # rechnungssumme = rechnungssumme + pos.art_total_price
                zwischensumme += pos.art_total_price

                startRowForPositions += 1

            for pos in currentRech.lAbzuege:
                actSheet.cell(row=startRowForPositions, column=2).value = pos.art_id
                actSheet.cell(row=startRowForPositions, column=3).value = pos.art_name
                actSheet.cell(row=startRowForPositions, column=4).value = pos.art_unit
                actSheet.cell(row=startRowForPositions, column=5).value = pos.art_price
                actSheet.cell(row=startRowForPositions, column=6).value = pos.art_amount
                actSheet.cell(row=startRowForPositions, column=7).value = pos.art_total_price * -1

                # rechnungssumme = rechnungssumme - pos.art_total_price
                zwischensumme -= pos.art_total_price

                startRowForPositions += 1

            for curCol in range(1, 8):
                actSheet.cell(row=startRowForPositions - 1, column=curCol).border = table_boarder_thick_bottom

            startRowForPositions += 1

            fillGreyBG: PatternFill = PatternFill(start_color='BFBFBF', end_color='BFBFBF',
                                                  fill_type="solid")  # used hex code for grey color

            if currentRech.bAddTaxes:
                actSheet.cell(row=startRowForPositions, column=1).value = "Zwischensumme"
                # UstBetrag = zwischensumme / 100 * self.taxRate
                # rechnungssumme = zwischensumme + UstBetrag
            else:
                actSheet.cell(row=startRowForPositions, column=1).value = "Rechnungssumme"
                # rechnungssumme = zwischensumme

            actSheet.cell(row=startRowForPositions, column=1).fill = fillGreyBG
            actSheet.cell(row=startRowForPositions, column=1).font = Font(bold=True)
            actSheet.cell(row=startRowForPositions, column=2).fill = fillGreyBG

            # Zwischensumme ist hier immer passend; nur wenn Taxes drauf sind ist es RechSum
            actSheet.cell(row=startRowForPositions, column=7).value = zwischensumme
            actSheet.cell(row=startRowForPositions, column=7).fill = fillGreyBG
            actSheet.cell(row=startRowForPositions, column=7).font = Font(bold=True)

            if currentRech.bAddTaxes:
                UstBetrag: float = zwischensumme / 100 * self.taxRate
                rechnungssumme: float = zwischensumme + UstBetrag

                startRowForPositions += 2
                tmpStringTaxRate: str = "+" + str(self.taxRate).replace(".0", "") + "% MwSt."
                actSheet.cell(row=startRowForPositions, column=5).value = tmpStringTaxRate
                actSheet.cell(row=startRowForPositions, column=5).font = Font(bold=True)
                actSheet.cell(row=startRowForPositions, column=7).value = UstBetrag
                actSheet.cell(row=startRowForPositions, column=7).font = Font(bold=True)
                actSheet.cell(row=startRowForPositions, column=7).number_format = '#,##0.00 €'

                startRowForPositions += 2

                for curCol in range(1, 8):
                    actSheet.cell(row=startRowForPositions - 1, column=curCol).border = table_boarder_thick_bottom
                startRowForPositions += 1

                actSheet.cell(row=startRowForPositions, column=1).value = "Rechnungssumme"
                actSheet.cell(row=startRowForPositions, column=1).fill = fillGreyBG
                actSheet.cell(row=startRowForPositions, column=1).font = Font(bold=True)
                actSheet.cell(row=startRowForPositions, column=2).fill = fillGreyBG

                actSheet.cell(row=startRowForPositions, column=7).value = rechnungssumme
                actSheet.cell(row=startRowForPositions, column=7).fill = fillGreyBG
                actSheet.cell(row=startRowForPositions, column=7).font = Font(bold=True)

            valuesSheet: worksheet = workbook["values"]
            valuesSheet["A2"] = currentRech.customer.cust_name
            valuesSheet["B2"] = zwischensumme  # für auswertung ist nur netto-summe relevant
            valuesSheet["C2"] = currentRech.strRechnungsNum
            valuesSheet["D2"] = currentRech.strAuftragsort
            valuesSheet["E2"] = currentRech.strRechnungsdatum

            workbook.save(filename=fPathSaveLocation)
            self.fPath_SavedExcel = fPathSaveLocation
            return True

        except Exception as exc:
            self.printExceptionDetails(exc, True)
            return False

    def loadCurrRechObjFromFile(self):
        fPath: str = os.path.join(os.path.normpath(self.lE_path2ConfigDir.text()), "latestCurrRechObj.obj")
        currRechObj = None
        fileToRead = None
        try:
            fileToRead = open(fPath, "rb")
            currRechObj = pickle.load(fileToRead)
            fileToRead.close()

        except Exception as exc:
            if fileToRead is not None:
                if not fileToRead.closed:
                    fileToRead.close()
            self.printExceptionDetails(exc, True)
            self.printInfoToLog("Fehler beim Lesen des RechObj-Files", True)
        return currRechObj

    def actionsAfterRechnungSave(self):
        self.incrementRechNum()
        self.setRechNumInGUI()
        self.writeCurrentRechNumToFile()
        self.writeRechYearToFile()
        self.resetDialog()

    def incrementRechNum(self):
        try:
            tmpRechnum: int = int(self.strRechNum)
            tmpRechnum += 1

            self.strRechNum = str(tmpRechnum)

        except Exception as exc:
            self.strRechNum = "1"
            self.printExceptionDetails(exc, True)

    def getCurrentRechnungDetails(self):
        currRechnung = CRechnung()
        currRechnung.lArticles = self.getArticles_FilledDetails()
        currRechnung.lAbzuege = self.getAbzuege_FilledDetails()

        currRechnung.strAuftragsort = self.tbAuftragsort.text()
        currRechnung.strAuftragsdatum = self.tbAuftragsdatum.text()
        currRechnung.strAuftragsnummer = self.tbAuftragsnummer.text()

        currRechnung.strRechnungsdatum = self.tbRechDatum.text()
        currRechnung.strRechnungsNum = self.tbRechNummer.text()
        currRechnung.customer = self.currentSelCustomer

        currRechnung.bAddTaxes = self.chk_AddTaxes.isChecked()

        return currRechnung

    def getArticles_FilledDetails(self):
        currRow: int = 0
        # TODO
        currentListArticles = []

        while currRow < self.twPositionen.rowCount():
            try:
                tmpArticle = CArticles()
                tmpArticle.art_id = self.twPositionen.item(currRow, 0).text()
                tmpArticle.art_name = self.twPositionen.item(currRow, 1).text()
                tmpArticle.art_unit = self.twPositionen.item(currRow, 2).text()

                str_art_price: str = self.twPositionen.item(currRow, 3).text()
                str_art_amount: str = self.twPositionen.item(currRow, 4).text()

                str_art_amount: str = str_art_amount.replace(",", ".")
                str_art_price: str = str_art_price.replace(",", ".")

                try:
                    tmpArticle.art_amount = float(str_art_amount)
                    tmpArticle.art_price = float(str_art_price)

                    if tmpArticle.art_amount > 0.0:
                        tmpArticle.art_total_price = tmpArticle.art_price * tmpArticle.art_amount
                        currentListArticles.append(tmpArticle)

                except Exception as exc:
                    self.printExceptionDetails(exc, True)

            except Exception as exc:
                if "NoneType" not in str(exc):
                    self.printExceptionDetails(exc, False)

            currRow += 1

        return currentListArticles

    def getAbzuege_FilledDetails(self):
        currRow: int = 0
        # todo
        currentListAbzuege = []

        while currRow < self.twAbzuege.rowCount():
            try:
                tmpArticle = CAbzuege()
                tmpArticle.art_id = self.twAbzuege.item(currRow, 0).text()
                tmpArticle.art_name = self.twAbzuege.item(currRow, 1).text()
                tmpArticle.art_unit = self.twAbzuege.item(currRow, 2).text()

                str_art_price: str = self.twAbzuege.item(currRow, 3).text()
                str_art_amount: str = self.twAbzuege.item(currRow, 4).text()

                str_art_amount = str_art_amount.replace(",", ".")
                str_art_price = str_art_price.replace(",", ".")

                try:
                    tmpArticle.art_amount = float(str_art_amount)
                    tmpArticle.art_price = float(str_art_price)

                    if tmpArticle.art_amount > 0.0:
                        tmpArticle.art_total_price = tmpArticle.art_price * tmpArticle.art_amount
                        currentListAbzuege.append(tmpArticle)

                except Exception as exc:
                    self.printExceptionDetails(exc, True)

            except Exception as exc:
                if "NoneType" not in str(exc):
                    self.printExceptionDetails(exc, False)

            currRow += 1

        return currentListAbzuege

    def printExceptionDetails(self, curException: Exception, bPrintToLog: bool):
        # curTimeStamp = str(datetime.datetime.now().strftime("<%H:%M:%S:%f>:\t"))
        # strOutput = curTimeStamp + "<ERROR>\t" + str(curException)
        strOutput: str = "<ERROR>\t" + str(curException)

        print(strOutput)

        if bPrintToLog:
            self.pTE_Logger.appendPlainText(strOutput)

    def printInfoToLog(self, strMsgToLog: str, bPrintToLog: bool):
        # curTimeStamp = str(datetime.datetime.now().strftime("<%H:%M:%S:%f>:\t"))
        # strOutput = curTimeStamp + "<INFO>\t" + strMsgToLog
        strOutput: str = "<INFO>\t" + strMsgToLog

        print(strOutput)

        if bPrintToLog:
            self.pTE_Logger.appendPlainText(strOutput)

    def checkIfYearChanged(self):
        strLatestYear: str = self.loadLatestYearFromFile()

        tmpYearOfToday: str = str(datetime.date.today().year)

        retYearSame: bool = strLatestYear != tmpYearOfToday

        return retYearSame

    def writeRechYearToFile(self):
        strRechYear: str = str((dateparser.parse(self.tbRechDatum.text(), settings={'DATE_ORDER': 'DMY'})).year)

        fPath: str = os.path.join(os.path.normpath(self.lE_path2ConfigDir.text()), self.lFName_LastRechYear.text())
        self.writeValueToFile(strRechYear, fPath)
